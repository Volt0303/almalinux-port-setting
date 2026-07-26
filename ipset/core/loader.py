"""loader.py - read CSV/xlsx supply files into normalized, validated rows.

Output model:
    PortRow   one LAN port's intended settings (one input line)
    Machine   one PC (serial number) with its list of PortRow
    LoadResult  machines + errors + warnings

Structural validation only (required columns/fields, duplicate IPs).
IP/netmask *format* validation is done in netmask.py (Step 4).

Python 3.9 compatible; CSV path is stdlib-only (openpyxl imported lazily).
"""
from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional

# Canonical field name -> accepted header aliases (all compared lower/stripped).
_ALIASES = {
    "sn":         ["sn", "serial", "serialno", "serial_no", "serialnumber", "シリアル", "シリアル番号"],
    "ifname":     ["ifname", "interface", "iface", "dev", "device"],
    "con_name":   ["con_name", "conname", "connection", "lan", "port", "lan_no", "lanno"],
    "ip_address": ["ip_address", "ip", "ipaddress", "ipaddr", "address"],
    "subnet":     ["subnet", "netmask", "mask", "subnetmask", "prefix"],
    "gateway":    ["gateway", "gw", "defaultgateway", "default_gateway"],
    "dns":        ["dns", "dns_server", "dnsserver", "nameserver"],
}
_REQUIRED_COLS = ["sn", "con_name", "ip_address", "subnet"]
_REQUIRED_VALS = ["sn", "con_name", "ip_address", "subnet"]


@dataclass
class PortRow:
    sn: str
    con_name: str
    ip_address: str
    subnet: str
    ifname: str = ""
    gateway: str = ""
    dns: str = ""
    source_line: int = 0          # 1-based line/row number in the input file


@dataclass
class Machine:
    sn: str
    ports: List[PortRow] = field(default_factory=list)


@dataclass
class LoadResult:
    machines: "Dict[str, Machine]" = field(default_factory=dict)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors

    def machine_list(self) -> List[Machine]:
        return list(self.machines.values())


def _build_header_map(header: List[str]) -> "Dict[str, int]":
    """Map canonical field -> column index, using aliases (case-insensitive)."""
    lookup = {}
    for idx, raw in enumerate(header):
        key = (raw or "").strip().lower()
        for canon, aliases in _ALIASES.items():
            if key in aliases and canon not in lookup:
                lookup[canon] = idx
    return lookup


def _rows_from_csv(path: str) -> List[List[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f)]


def _rows_from_xlsx(path: str) -> List[List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required to read .xlsx files (pip install openpyxl)"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(["" if c is None else str(c).strip() for c in row])
    wb.close()
    return out


def load(path: str) -> LoadResult:
    """Load a CSV/xlsx supply file into a LoadResult."""
    res = LoadResult()
    if not os.path.isfile(path):
        res.errors.append("File not found: %s" % path)
        return res

    ext = os.path.splitext(path)[1].lower()
    try:
        rows = _rows_from_xlsx(path) if ext in (".xlsx", ".xlsm") else _rows_from_csv(path)
    except Exception as e:  # noqa: BLE001 - surface any parse failure as an error
        res.errors.append("Failed to read %s: %s" % (path, e))
        return res

    # Drop fully-blank leading rows, find the header.
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        res.errors.append("File is empty: %s" % path)
        return res

    header = [(c or "").strip() for c in rows[0]]
    hmap = _build_header_map(header)
    missing = [c for c in _REQUIRED_COLS if c not in hmap]
    if missing:
        res.errors.append(
            "Missing required column(s): %s (found header: %s)"
            % (", ".join(missing), ", ".join(header))
        )
        return res

    def cell(row: List[str], canon: str) -> str:
        idx = hmap.get(canon)
        if idx is None or idx >= len(row):
            return ""
        return (row[idx] or "").strip()

    for lineno, row in enumerate(rows[1:], start=2):  # header is line 1
        rec = {c: cell(row, c) for c in _ALIASES}
        empty_required = [c for c in _REQUIRED_VALS if not rec[c]]
        if empty_required:
            res.errors.append(
                "Line %d: missing value for %s" % (lineno, ", ".join(empty_required))
            )
            continue
        pr = PortRow(
            sn=rec["sn"], con_name=rec["con_name"], ip_address=rec["ip_address"],
            subnet=rec["subnet"], ifname=rec["ifname"], gateway=rec["gateway"],
            dns=rec["dns"], source_line=lineno,
        )
        res.machines.setdefault(pr.sn, Machine(sn=pr.sn)).ports.append(pr)

    _check_duplicate_ips(res)
    return res


def _check_duplicate_ips(res: LoadResult) -> None:
    """Warn on IPs repeated within a single machine (e.g. sample row 19)."""
    for m in res.machines.values():
        seen: Dict[str, PortRow] = {}
        for p in m.ports:
            if p.ip_address in seen:
                first = seen[p.ip_address]
                res.warnings.append(
                    "SN %s: duplicate IP %s on %s (line %d) and %s (line %d)"
                    % (m.sn, p.ip_address, first.con_name, first.source_line,
                       p.con_name, p.source_line)
                )
            else:
                seen[p.ip_address] = p


def find_machine(res: LoadResult, sn: str) -> Optional[Machine]:
    return res.machines.get(sn)
